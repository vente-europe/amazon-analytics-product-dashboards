"""Quick inspector to check XLSX structure."""
from openpyxl import load_workbook
from pathlib import Path

reviews_dir = Path(__file__).resolve().parent.parent  # reviews/DE
files = [
    'B00AE6XK3O-DE-Reviews-20260429.xlsx',
    'B0CPLQRWBG-DE-Reviews-20260429.xlsx',
    'B0CT8QSYT5-DE-Reviews-20260429.xlsx',
    'B0DS8SK65N-DE-Reviews-20260429.xlsx',
    'B0GDDRCP6M-DE-Reviews-20260429.xlsx',
]
for f in files:
    p = reviews_dir / f
    if not p.exists():
        print(f"MISSING: {p}")
        continue
    wb = load_workbook(p, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    print(f"=== {f} ===")
    print(f"Rows: {len(rows)}")
    if rows:
        print(f"Headers: {rows[0]}")
        if len(rows) > 1:
            print(f"Sample: {rows[1]}")
    wb.close()
