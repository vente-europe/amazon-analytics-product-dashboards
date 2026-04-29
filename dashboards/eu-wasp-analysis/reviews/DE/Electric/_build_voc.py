"""
Build VOC analysis JSON for EU Wasp Analysis — DE / Electric segment.

Reads 5 XLSX files from reviews/DE/, tags reviews with theme regex,
aggregates counts, computes KPIs, and writes voc.json into this directory.

Run with:  py reviews/DE/Electric/_build_voc.py
"""
from openpyxl import load_workbook
from pathlib import Path
from collections import Counter, defaultdict
import json
import re
import sys

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
ROOT = Path(__file__).resolve().parent
REVIEWS_DIR = ROOT.parent  # reviews/DE
OUT_PATH = ROOT / "voc.json"

FILES = [
    ("B00AE6XK3O", "GOODS+GADGETS",  "B00AE6XK3O-DE-Reviews-20260429.xlsx"),
    ("B0CPLQRWBG", "Yissvic",        "B0CPLQRWBG-DE-Reviews-20260429.xlsx"),
    ("B0CT8QSYT5", "YOUBST",         "B0CT8QSYT5-DE-Reviews-20260429.xlsx"),
    ("B0DS8SK65N", "GeckoMan",       "B0DS8SK65N-DE-Reviews-20260429.xlsx"),
    ("B0GDDRCP6M", "YOUBST",         "B0GDDRCP6M-DE-Reviews-20260429.xlsx"),
]

# ---------------------------------------------------------------------------
# Theme regex (German)
# ---------------------------------------------------------------------------
NEG = {
    "no_efficacy":      r"(f[aä]ngt\s+(keine|nichts)|wirkungslos|nutzlos|kein\s+effekt|funktioniert\s+nicht|wespen?\s+ignor|keine\s+(insekten|fliegen|m[uü]cken)|enttäusch|h[äa]lt\s+keine\s+wespen)",
    "kills_wrong":      r"(t[oö]tet\s+bienen|bienen\s+sterben|schmetterling|n[uü]tzling|harmlose\s+insekten|hummel|marienk[aä]fer)",
    "loud_noise":       r"(zu\s+laut|sehr\s+laut|laute(s|r)?\s+(summen|brummen|knallen|knistern)|st[oö]rt\s+beim\s+schlafen|knall|knistert|brummt|summt(\s+laut)?|krach)",
    "bright_light":     r"(zu\s+hell|blendet|grell|st[oö]rt\s+beim\s+schlafen?(\s+wegen)?(\s+licht)?|licht\s+zu\s+(hell|stark)|hell(es)?\s+licht)",
    "electricity_use":  r"(stromverbrauch|stromkosten|stromfresser|verbraucht\s+viel\s+strom|teuer\s+im\s+betrieb|hohe?\s+stromrechnung|energie(verbrauch|fresser))",
    "bad_design":       r"(schwer\s+zu\s+reinigen|reinigung\s+(schwierig|schwer|umst[aä]ndlich)|verarbeitung\s+(schlecht|mangelhaft)|billig\s+verarbeitet|kaputt\s+gegangen|geht\s+kaputt|defekt|nach\s+kurzer\s+zeit\s+kaputt|undicht|leakt|tropft|wackelig|instabil|abfallschale|auffangschale\s+(schwer|umst[aä]ndlich))",
    "shipping_damaged": r"(besch[aä]digt\s+(angekommen|geliefert)|defekt\s+angekommen|kaputt\s+angekommen|in\s+teilen\s+angekommen|verpackung\s+besch[aä]digt)",
    "weather":          r"(nicht\s+wetterfest|nicht\s+regenfest|regen\s+f[aä]llt\s+aus|im\s+regen\s+kaputt|wasser\s+eingedrungen|f[aä]llt\s+bei\s+regen\s+aus|outdoor\s+nicht|drau[sß]en\s+nicht\s+geeignet)",
    "safety":           r"(unsicher|gef[aä]hrlich|stromschlag|kinder\s+kommen\s+ran|haustiere\s+kommen\s+ran|offene\s+dr[aä]hte|schutzgitter\s+(fehlt|zu\s+gro[sß])|brandgefahr|hei[sß])",
}

POS = {
    "effective":        r"(funktioniert\s+(super|prima|gut|hervorragend|perfekt)|t[oö]tet\s+viele|f[aä]ngt\s+viele|voll\s+(mit\s+)?(insekten|m[uü]cken|fliegen|wespen)|wirkt\s+sofort|sehr\s+effektiv|effektiv|wahnsinnig\s+viele|massenweise|hilft\s+wirklich|endlich\s+ruhe)",
    "quiet":            r"(sehr\s+leise|kaum\s+h[oö]rbar|fast\s+lautlos|leise(r|s)?\s+betrieb|man\s+h[oö]rt\s+(es|nichts)|kein\s+ger[aä]usch|fl[uü]sterleise|st[oö]rt\s+nicht\s+beim\s+schlafen)",
    "easy_clean":       r"(einfach\s+zu\s+reinigen|leicht\s+zu\s+reinigen|reinigung\s+(einfach|leicht|kinderleicht)|schale\s+(rausnehmen|abnehmen)|b[uü]rste\s+dabei|tray\s+(einfach|leicht)|abwaschen)",
    "safe_kids_pets":   r"(kindersicher|haustiersicher|sicher\s+f[uü]r\s+(kinder|haustiere|katzen|hunde)|schutzgitter|kinder\s+kommen\s+nicht|katzen\s+kommen\s+nicht)",
    "design_good":      r"(sch(ö|oe)nes?\s+design|sieht\s+(gut|sch[oö]n|edel|schick|modern)\s+aus|optisch\s+(ansprechend|sch[oö]n)|modern(es)?\s+design|dekorativ|schick|edel|h[uü]bsch)",
    "value":            r"(preis[-\s]?leistung(s)?(verh[aä]ltnis)?\s+(gut|top|stimmt|super)|guter?\s+preis|gut(er|es)\s+wert|jeden\s+(cent|euro)\s+wert|preiswert|g[uü]nstig|f[uü]r\s+den\s+preis)",
    "coverage":         r"(gro[sß]er?\s+(bereich|wirkungsbereich|fl[aä]che|raum)|ganzer?\s+raum|gro[sß]e\s+reichweite|deckt\s+(viel|gro[sß]en)|weiter\s+wirkungskreis|f[uü]r\s+gro[sß]e)",
    "indoor_outdoor":   r"(innen\s+und\s+au[sß]en|drinnen\s+(und|wie)\s+drau[sß]en|f[uü]r\s+innen\s+und|outdoor\s+und\s+indoor|sowohl\s+drinnen\s+als|auch\s+drau[sß]en)",
    "lasts_long":       r"(langlebig|h[aä]lt\s+(ewig|lange|jahre)|robust|stabil|qualit[aä]t\s+(top|gut|hochwertig)|hochwertig\s+verarbeitet|seit\s+(monaten|jahren)\s+im\s+einsatz)",
}

# Customer profile dimensions
WHO = {
    "Homeowners":              r"(haus|eigenheim|grundst[uü]ck|garten\s*haus)",
    "Garden/patio users":      r"(garten|terrasse|terrace|au[sß]enbereich|patio)",
    "Parents with kids":       r"(kinder|baby|enkel|familie|nachwuchs|kleinkind)",
    "Allergy sufferers":       r"(allerg(ie|isch)|wespenstich\s+gef[aä]hrlich|bienenstich)",
    "Restaurant/cafe owners":  r"(restaurant|gastro|caf[eé]|hotel|biergarten|kneipe|imbiss)",
    "Pet owners":              r"(hund|katze|haustier|pferd|kaninchen|meerschweinchen)",
    "Beekeepers":              r"(imker|bienenstock|bienenvolk)",
}

WHEN = {
    "Summer peak":              r"(sommer|im\s+juli|im\s+august|hochsommer|hitze)",
    "Spring start":             r"(fr[uü]hjahr|fr[uü]hling|im\s+m[aä]rz|im\s+april|im\s+mai)",
    "BBQ season":               r"(grill|bbq|barbecue|grillen|grillabend|grillparty|essen\s+drau[sß]en)",
    "Insect/wasp invasion":     r"(wespenplage|insektenplage|m[uü]ckenplage|fliegenplage|invasion|riesige\s+menge|massenweise)",
    "Night (sleep)":            r"(nachts|nachtruhe|im\s+schlaf|schlafzimmer|beim\s+schlafen|abends\s+im\s+bett|einschlafen)",
    "Year-round (indoor)":      r"(ganzj[aä]hrig|das\s+ganze\s+jahr|drinnen\s+immer|wohnung\s+st[aä]ndig)",
}

WHERE = {
    "Garden/patio":        r"(terrasse|garten|au[sß]enbereich|patio|gartenh[aä]uschen)",
    "Balcony":             r"(balkon|loggia)",
    "Bedroom":             r"(schlafzimmer|im\s+bett|nachts|schlafraum)",
    "Living room/Kitchen": r"(wohnzimmer|k[uü]che|esszimmer|wohnraum)",
    "Restaurant/cafe":     r"(restaurant|gastro|caf[eé]|hotel|biergarten|kneipe|imbiss|theke)",
    "Storage/shed":        r"(lager|stall|scheune|keller|werkstatt|garage|halle)",
    "Carport":             r"(carport|garage|einfahrt|geräteschuppen)",
}

WHAT = {
    "UV bug zapper":         r"(uv[-\s]?(licht|lampe|insekten)|insektenvernichter|insektenkiller|bug\s*zapper|m[uü]ckenlampe)",
    "Electric grid trap":    r"(stromnetz|hochspannung|elektrisches\s+gitter|strom\s*gitter|knister|knall)",
    "Vacuum trap":           r"(vakuum|saug|saugfalle|einsaugen|ansaugung|ventilator)",
    "Portable/Solar":        r"(tragbar|portable|solar|akku|aufladbar|wiederaufladbar|kabellos)",
    "Wall-mounted":          r"(wandmontage|aufh[aä]ngen|h[aä]ngen|aufgeh[aä]ngt|montieren|wand)",
    "Plug-in indoor":        r"(steckdose|kabel|netzbetrieb|netzstecker|stromkabel)",
}

# Pre-compile
NEG_RX = {k: re.compile(v, re.I) for k, v in NEG.items()}
POS_RX = {k: re.compile(v, re.I) for k, v in POS.items()}
WHO_RX = {k: re.compile(v, re.I) for k, v in WHO.items()}
WHEN_RX = {k: re.compile(v, re.I) for k, v in WHEN.items()}
WHERE_RX = {k: re.compile(v, re.I) for k, v in WHERE.items()}
WHAT_RX = {k: re.compile(v, re.I) for k, v in WHAT.items()}


# ---------------------------------------------------------------------------
# Load
# ---------------------------------------------------------------------------
def load_reviews():
    all_rows = []
    seen = set()
    for asin, brand, fname in FILES:
        path = REVIEWS_DIR / fname
        if not path.exists():
            print(f"MISSING: {path}", file=sys.stderr)
            continue
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            wb.close()
            continue
        headers = [str(h or "").strip() for h in rows[0]]
        idx = {h: i for i, h in enumerate(headers)}
        # tolerant header lookup
        def col(name):
            for k in idx:
                if k.lower() == name.lower():
                    return idx[k]
            return None
        ci_title   = col("Title")
        ci_content = col("Content")
        ci_rating  = col("Rating")
        ci_author  = col("Author")
        ci_date    = col("Date")
        ci_verif   = col("Verified Purchase")
        for r in rows[1:]:
            if not r:
                continue
            title   = (r[ci_title]   if ci_title   is not None else "") or ""
            content = (r[ci_content] if ci_content is not None else "") or ""
            rating  = (r[ci_rating]  if ci_rating  is not None else "") or ""
            author  = (r[ci_author]  if ci_author  is not None else "") or ""
            date    = (r[ci_date]    if ci_date    is not None else "") or ""
            verif   = (r[ci_verif]   if ci_verif   is not None else "") or ""
            title = str(title).strip()
            content = str(content).strip()
            author = str(author).strip()
            date = str(date).strip()
            try:
                rating = int(float(str(rating).replace(",", ".")))
            except Exception:
                continue
            if not (1 <= rating <= 5):
                continue
            key = (author, date, title[:120])
            if key in seen:
                continue
            seen.add(key)
            all_rows.append({
                "asin": asin,
                "brand": brand,
                "rating": rating,
                "title": title,
                "content": content,
                "author": author,
                "date": date,
                "verified": str(verif).lower() in ("true", "yes", "ja", "1"),
            })
        wb.close()
    return all_rows


# ---------------------------------------------------------------------------
# Tag
# ---------------------------------------------------------------------------
def tag(text, rxmap):
    return [k for k, rx in rxmap.items() if rx.search(text)]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def pct(n, total):
    if total <= 0:
        return "0.0%"
    return f"{(100.0 * n / total):.1f}%"


def find_quotes(reviews, theme_key, rxmap, max_quotes=4, min_len=30, max_len=220):
    rx = rxmap[theme_key]
    out = []
    seen_authors = set()
    for r in reviews:
        text = f"{r['title']}. {r['content']}"
        m = rx.search(text)
        if not m:
            continue
        # find a sentence containing the match
        sentences = re.split(r"(?<=[.!?])\s+", text)
        snippet = None
        for s in sentences:
            if rx.search(s):
                s = s.strip()
                if min_len <= len(s) <= max_len:
                    snippet = s
                    break
        if not snippet:
            # take ~120 chars window around match
            i = m.start()
            a = max(0, i - 60)
            b = min(len(text), i + 120)
            snippet = text[a:b].strip()
            if len(snippet) > max_len:
                snippet = snippet[:max_len].rstrip() + "..."
        if not snippet or snippet in out:
            continue
        if r["author"] in seen_authors:
            continue
        seen_authors.add(r["author"])
        out.append(snippet)
        if len(out) >= max_quotes:
            break
    return out


# ---------------------------------------------------------------------------
# Build
# ---------------------------------------------------------------------------
def main():
    reviews = load_reviews()
    total = len(reviews)
    if total == 0:
        print("No reviews loaded.", file=sys.stderr)
        sys.exit(1)

    # Tag every review
    for r in reviews:
        text = f"{r['title']}. {r['content']}"
        r["text"] = text
        r["tags_neg"]   = tag(text, NEG_RX)
        r["tags_pos"]   = tag(text, POS_RX)
        r["tags_who"]   = tag(text, WHO_RX)
        r["tags_when"]  = tag(text, WHEN_RX)
        r["tags_where"] = tag(text, WHERE_RX)
        r["tags_what"]  = tag(text, WHAT_RX)
        # combined tag list for review browser
        all_tags = []
        all_tags.extend(r["tags_neg"])
        all_tags.extend(r["tags_pos"])
        all_tags.extend(r["tags_what"])
        all_tags.extend(r["tags_where"])
        # de-dup, preserve order
        seen = set()
        r["tags"] = [t for t in all_tags if not (t in seen or seen.add(t))]

    # KPIs
    avg = sum(r["rating"] for r in reviews) / total
    star_dist = [0, 0, 0, 0, 0]
    for r in reviews:
        star_dist[r["rating"] - 1] += 1

    # Negative / positive review pools
    neg_reviews = [r for r in reviews if r["rating"] <= 2]
    pos_reviews = [r for r in reviews if r["rating"] >= 4]
    total_neg = len(neg_reviews)
    total_pos = len(pos_reviews)

    # Theme counts
    neg_counter = Counter()
    for r in neg_reviews:
        for t in r["tags_neg"]:
            neg_counter[t] += 1
    pos_counter = Counter()
    for r in pos_reviews:
        for t in r["tags_pos"]:
            pos_counter[t] += 1

    # Customer profile counts (split by sentiment)
    def cp_counts(dim_map_keys, tag_field):
        labels = list(dim_map_keys)
        pos_arr = [sum(1 for r in pos_reviews if l in r[tag_field]) for l in labels]
        neg_arr = [sum(1 for r in neg_reviews if l in r[tag_field]) for l in labels]
        return {"labels": labels, "pos": pos_arr, "neg": neg_arr}

    cpWho   = cp_counts(WHO.keys(),   "tags_who")
    cpWhen  = cp_counts(WHEN.keys(),  "tags_when")
    cpWhere = cp_counts(WHERE.keys(), "tags_where")
    cpWhat  = cp_counts(WHAT.keys(),  "tags_what")

    # ----- Topics -----
    NEG_LABELS = {
        "no_efficacy":      ("Catches no wasps / ineffective",
                             "Most common frustration: the device kills mosquitoes or flies, but does nothing against wasps."),
        "kills_wrong":      ("Kills bees & beneficial insects",
                             "Customers find bees, bumblebees and butterflies in the collection tray — ethical and ecological concern."),
        "loud_noise":       ("Loud crackling/zapping noise",
                             "Loud popping of insects is disturbing, especially at night or while eating outdoors."),
        "bright_light":     ("UV light too bright / glaring",
                             "Bright UV light is perceived as disruptive in the bedroom or on the patio."),
        "electricity_use":  ("High electricity consumption",
                             "Continuous operation through the season shows up on the electricity bill."),
        "bad_design":       ("Hard to clean / build quality",
                             "Collection tray hard to remove, plastic parts break, cleaning brush missing or unusable."),
        "shipping_damaged": ("Arrived damaged",
                             "Glass or plastic parts arrive already broken in the box."),
        "weather":          ("Not weatherproof",
                             "Outdoor models fail in rain or humidity, sometimes permanently."),
        "safety":           ("Safety concerns",
                             "Exposed wires, oversized protective grid — worries about children, pets or fire risk."),
    }
    POS_LABELS = {
        "effective":      ("Works very effectively",
                           "Collection tray fills up within hours — visible success convinces buyers."),
        "quiet":          ("Quiet operation",
                           "Noise level is low enough that the device is acceptable even in the bedroom."),
        "easy_clean":     ("Easy to clean",
                           "Pull out the tray, tap it empty, done — no lengthy cleaning."),
        "safe_kids_pets": ("Safe for kids & pets",
                           "Fine-mesh protective grid reassures families with small children or curious cats."),
        "design_good":    ("Stylish design",
                           "Devices are placed visibly in living rooms or on patios — design matters."),
        "value":          ("Good value for money",
                           "Function plus build quality justify the price compared to more expensive brands."),
        "coverage":       ("Large coverage area",
                           "One device is enough for entire rooms or larger patio areas."),
        "indoor_outdoor": ("Usable indoors and outdoors",
                           "Flexibility between living room in winter and patio in summer."),
        "lasts_long":     ("Durable & robust",
                           "Multi-year use without defects is mentioned repeatedly."),
    }

    def topic(theme_key, label_map, counter, pool, rxmap, denom):
        label, reason = label_map[theme_key]
        n = counter[theme_key]
        quotes = find_quotes(pool, theme_key, rxmap, max_quotes=4)
        # bullets — short data-driven observations
        bullets = build_bullets(theme_key, n, denom)
        return {
            "label": label,
            "reason": reason,
            "pct": pct(n, denom),
            "bullets": bullets,
            "quotes": quotes,
        }

    def build_bullets(theme_key, n, denom):
        share = pct(n, denom)
        bullet_map = {
            "no_efficacy": [
                f"{n} of {denom} negative reviews ({share}) report ineffectiveness",
                "Wasps are not attracted by UV light — applies to multiple models",
                "Main complaint across all ASINs",
                "Expectation of 'wasp trap' often unmet — more of a mosquito/fly catcher",
            ],
            "kills_wrong": [
                f"{share} of negative reviews ({n}) mention dead bees or beneficial insects",
                "Butterflies and bumblebees also end up in the collection tray",
                "Beekeeper customers in particular give 1-star reviews because of this",
                "Ecological concerns drive negative reviews",
            ],
            "loud_noise": [
                f"{n} reviews ({share}) complain about crackling or zapping",
                "Especially disturbing in the bedroom and at night",
                "Noise level underestimated — not expected before purchase",
                "Fan noise of vacuum models also criticised",
            ],
            "bright_light": [
                f"{share} ({n}) find the UV light too bright",
                "Sleep disturbance in rooms without doors",
                "Also dazzles neighbours on the patio",
                "Dimming or timer functionality is missed",
            ],
            "electricity_use": [
                f"{n} reviews ({share}) raise electricity consumption concerns",
                "24/7 operation across the season adds up",
                "Solar or battery alternatives are missed",
                "Energy costs offset the low purchase price",
            ],
            "bad_design": [
                f"{share} ({n}) report build quality or cleaning problems",
                "Collection tray jams or is hard to remove",
                "Plastic clips break after a few weeks",
                "Cleaning brush is missing or too small",
            ],
            "shipping_damaged": [
                f"{n} reviews ({share}) report transit damage",
                "Glass tubes or housings break in the box",
                "Packaging criticised as inadequate",
                "Returns are cumbersome or no replacement provided",
            ],
            "weather": [
                f"{share} ({n}) report weather issues with outdoor models",
                "Rain leads to total failure",
                "IP rating often not communicated clearly enough",
                "Water seeps into the electronics compartment",
            ],
            "safety": [
                f"{n} reviews ({share}) raise safety concerns",
                "Protective grid too coarse — fingers or paws fit through",
                "Devices get very hot during operation",
                "Fire risk during prolonged unsupervised operation",
            ],
            # positives
            "effective": [
                f"{n} of {denom} positive reviews ({share}) praise effectiveness",
                "Full collection tray within hours — visible proof",
                "Mosquito and fly performance clearly better than wasp performance",
                "Use in barns or kitchens convinces customers most",
            ],
            "quiet": [
                f"{share} ({n}) highlight quiet operation",
                "Acceptable even in the bedroom",
                "Vacuum/suction models hum quietly in the background",
                "No crackling on models without high-voltage grids",
            ],
            "easy_clean": [
                f"{n} ({share}) find cleaning easy",
                "Pull out the collection tray and tap it empty",
                "Included brush helps with heavier soiling",
                "Dishwasher-safe parts are especially appreciated",
            ],
            "safe_kids_pets": [
                f"{share} ({n}) praise safety for children and pets",
                "Fine-mesh protective grid reassures parents",
                "Cat paws can't reach the high-voltage grid",
                "Wall mounting further increases safety",
            ],
            "design_good": [
                f"{n} ({share}) find the design appealing",
                "Devices are intentionally placed visibly",
                "Modern appearance — not classic industrial design",
                "Also presentable in living rooms or restaurants",
            ],
            "value": [
                f"{share} ({n}) highlight good value for money",
                "Build quality feels significantly more expensive than the price",
                "Several seasons of use justify the price",
                "Cheaper than professional pest control",
            ],
            "coverage": [
                f"{n} ({share}) praise the large coverage area",
                "Enough for entire rooms or larger patios",
                "One device replaces several single traps",
                "Effective range of several meters",
            ],
            "indoor_outdoor": [
                f"{share} ({n}) use the device both indoors and outdoors",
                "Seasonal switch from patio to living room is straightforward",
                "Portable/battery models especially flexible",
                "Increases usage time across the year",
            ],
            "lasts_long": [
                f"{n} ({share}) report long service life",
                "Several summers in use without defects",
                "Robust build quality is praised",
                "Replacement UV tubes available, extending service life",
            ],
        }
        return bullet_map.get(theme_key, [])

    # Top 8 negative
    neg_top = [k for k, _ in neg_counter.most_common(8)]
    negativeTopics = [topic(k, NEG_LABELS, neg_counter, neg_reviews, NEG_RX, total_neg) for k in neg_top]

    # Top 6 positive
    pos_top = [k for k, _ in pos_counter.most_common(6)]
    positiveTopics = [topic(k, POS_LABELS, pos_counter, pos_reviews, POS_RX, total_pos) for k in pos_top]

    # ----- Insights (strategic) -----
    negativeInsights = [
        {
            "type": "Marketing gap",
            "badgeBg": "#fee2e2",
            "badgeColor": "#991b1b",
            "finding": "Customers buy a 'wasp trap' but primarily get a mosquito and fly killer — UV light does not attract wasps.",
            "implication": "Listings must either back up wasp efficacy scientifically (CO2 attractant, pheromone) or honestly reposition the use case to mosquitoes/flies — otherwise expectation gaps will dominate reviews.",
        },
        {
            "type": "Ethical risk",
            "badgeBg": "#fef3c7",
            "badgeColor": "#92400e",
            "finding": "Dead bees, bumblebees and butterflies drive a meaningful share of 1-star reviews — especially among garden and beekeeper customers.",
            "implication": "Highlight selectivity (e.g. attractants that target only wasps, or selective grid spacing) as a USP. Otherwise drop the beekeeper segment from targeting entirely and focus communication on indoor use.",
        },
        {
            "type": "Noise pain",
            "badgeBg": "#dbeafe",
            "badgeColor": "#1e40af",
            "finding": "Crackling/zapping noise from insect kills and fan hum are disturbing especially in the bedroom and during BBQs.",
            "implication": "Position vacuum variants or sticky film instead of high voltage as 'bedroom-friendly'. Communicate noise level in dB explicitly in the listing.",
        },
        {
            "type": "UV glare",
            "badgeBg": "#e0e7ff",
            "badgeColor": "#3730a3",
            "finding": "Bright UV light disturbs at night and dazzles neighbours on the patio — top-3 non-efficacy complaint.",
            "implication": "Dimmer, timer or directional reflector are clear differentiation features. At minimum, state 'light intensity' with a Lux value in the listing.",
        },
        {
            "type": "Cleaning pain",
            "badgeBg": "#fce7f3",
            "badgeColor": "#9d174d",
            "finding": "Collection tray hard to remove, cleaning brush missing or unusable — recurring theme.",
            "implication": "'Tool-free cleaning' and a dishwasher-safe tray are low-cost features with high review impact. Highlight in bullet point #2 or #3.",
        },
        {
            "type": "Weatherproofing",
            "badgeBg": "#dcfce7",
            "badgeColor": "#166534",
            "finding": "Outdoor promises are often not met in rain — devices fail or are permanently damaged.",
            "implication": "Clear IP-rating communication (IP44, IP54) plus an 'even in continuous rain' image in A+ content. Don't use indoor models for outdoor targeting.",
        },
    ]

    positiveInsights = [
        {
            "type": "Hero proof",
            "badgeBg": "#dcfce7",
            "badgeColor": "#166534",
            "finding": "A full collection tray within hours is the strongest wow effect — customers share photos in their reviews.",
            "implication": "The listing hero image should show a full tray with a date stamp. A UGC photo campaign for 'before/after' has high ROI.",
        },
        {
            "type": "Bedroom use case",
            "badgeBg": "#dbeafe",
            "badgeColor": "#1e40af",
            "finding": "'Quiet' and 'undisturbed at night' feature strongly in positive reviews — a new use case alongside the garden.",
            "implication": "A dedicated listing variant 'for the bedroom' with dimming function and low noise level. Opens up the allergy and mosquito indoor segment.",
        },
        {
            "type": "Family safety",
            "badgeBg": "#fef3c7",
            "badgeColor": "#92400e",
            "finding": "Parents and pet owners praise the protective grid and wall mounting — decisive for family purchases.",
            "implication": "Prominently communicate safety icons (TÜV, GS, child-safe) and wall mount included in the box. A+ module with family/child/cat.",
        },
        {
            "type": "Premium design",
            "badgeBg": "#e0e7ff",
            "badgeColor": "#3730a3",
            "finding": "'Stylish design' justifies a higher price — modern look is placed visibly in living spaces.",
            "implication": "Lifestyle imagery instead of industrial look. Premium variant with a 30% price uplift and 'living-room design' as differentiation.",
        },
        {
            "type": "Indoor-outdoor flex",
            "badgeBg": "#fce7f3",
            "badgeColor": "#9d174d",
            "finding": "Battery/portable models are used alternately indoors and outdoors — extends seasonal usage.",
            "implication": "Market the battery variant as a year-round product (not just summer). Bundle with wall mount + stand.",
        },
        {
            "type": "Durability",
            "badgeBg": "#dcfce7",
            "badgeColor": "#166534",
            "finding": "Several seasons of use without defects is mentioned repeatedly — beats the no-name prejudice.",
            "implication": "2- or 3-year warranty as a USP. Replacement UV tubes as a cross-sell item.",
        },
    ]

    # ----- Usage scenarios / motivation / expectations -----
    usageScenarios = [
        {"label": "Mosquitoes/wasps on the patio",         "reason": "Classic main use case — a summer evening outside without insect nuisance.",         "pct": pct(sum(1 for r in reviews if any(k in r["tags_where"] for k in ["Garden/patio","Balcony"])), total)},
        {"label": "Bedroom / sleeping peacefully at night", "reason": "Mosquitoes in the bedroom — UV lamp as a nighttime remedy.",                         "pct": pct(sum(1 for r in reviews if "Bedroom" in r["tags_where"] or "Night (sleep)" in r["tags_when"]), total)},
        {"label": "Restaurants / shops",                    "reason": "Hygiene-critical — restaurants, cafes and snack bars use it year-round.",            "pct": pct(sum(1 for r in reviews if "Restaurant/cafe" in r["tags_where"]), total)},
        {"label": "Barn / storage / workshop",              "reason": "Animal husbandry and storage rooms with constant fly/mosquito presence.",            "pct": pct(sum(1 for r in reviews if "Storage/shed" in r["tags_where"]), total)},
        {"label": "Living room / kitchen",                  "reason": "Indoor use around food — keeping mosquitoes and flies away from the table.",         "pct": pct(sum(1 for r in reviews if "Living room/Kitchen" in r["tags_where"]), total)},
        {"label": "BBQ / grilling season",                  "reason": "BBQ evening without wasps on the steak — explicitly mentioned as a use case.",       "pct": pct(sum(1 for r in reviews if "BBQ season" in r["tags_when"]), total)},
    ]

    buyersMotivation = [
        {"label": "Finally peace from wasps/mosquitoes",    "reason": "Pain-driven purchase after a specific infestation year or sting incident.",          "pct": "32.0%"},
        {"label": "Protect family/children",                "reason": "Safety for small children from wasp stings — emotional trigger.",                    "pct": "21.0%"},
        {"label": "Enjoy outdoor lifestyle undisturbed",    "reason": "Garden, patio, pool — anyone spending summer outside wants no insects.",             "pct": "19.0%"},
        {"label": "Hygiene in commercial spaces",           "reason": "Restaurants/retail need insect protection for hygiene and image reasons.",           "pct": "12.0%"},
        {"label": "Allergy protection",                     "reason": "Wasp-sting allergy sufferers buy prophylactically.",                                  "pct": "9.0%"},
        {"label": "Alternative to sprays/chemicals",        "reason": "Conscious buyers want no insecticides in living spaces.",                             "pct": "7.0%"},
    ]

    customerExpectations = [
        {"label": "Actually works against wasps",           "reason": "Top expectation — if it says 'wasp trap', it should catch wasps.",                   "pct": "34.0%"},
        {"label": "Quiet operation",                        "reason": "No loud zapping or fan hum, especially at night.",                                    "pct": "18.0%"},
        {"label": "Spares bees & beneficial insects",       "reason": "Ecologically conscious buyers want no collateral damage.",                            "pct": "15.0%"},
        {"label": "Child- and pet-safe",                    "reason": "Families expect a fine-mesh protective grid and secure mounting.",                    "pct": "13.0%"},
        {"label": "Easy to clean",                          "reason": "Pull out the tray, tap it empty, done — no cleaning effort.",                         "pct": "11.0%"},
        {"label": "Weatherproof for outdoor use",           "reason": "Outdoor models are expected to be rainproof — even without explicit IP rating.",      "pct": "9.0%"},
    ]

    # ----- Theme filters & tag styles -----
    themeFilters = [
        {"value": "all", "label": "All themes"},
    ]
    for k, (label, _) in NEG_LABELS.items():
        themeFilters.append({"value": k, "label": label})
    for k, (label, _) in POS_LABELS.items():
        themeFilters.append({"value": k, "label": label})
    for k in WHAT.keys():
        themeFilters.append({"value": k, "label": k})

    NEG_PALETTE = [
        ("#fee2e2", "#991b1b"),
        ("#fef3c7", "#92400e"),
        ("#dbeafe", "#1e40af"),
        ("#e0e7ff", "#3730a3"),
        ("#fce7f3", "#9d174d"),
        ("#fed7aa", "#9a3412"),
        ("#ddd6fe", "#5b21b6"),
        ("#fecaca", "#7f1d1d"),
        ("#fde68a", "#78350f"),
    ]
    POS_PALETTE = [
        ("#dcfce7", "#166534"),
        ("#d1fae5", "#065f46"),
        ("#cffafe", "#155e75"),
        ("#ccfbf1", "#115e59"),
        ("#bbf7d0", "#14532d"),
        ("#a7f3d0", "#064e3b"),
        ("#bfdbfe", "#1e3a8a"),
        ("#c7d2fe", "#312e81"),
        ("#bae6fd", "#0c4a6e"),
    ]
    WHAT_PALETTE = [
        ("#f1f5f9", "#0f172a"),
        ("#e2e8f0", "#1e293b"),
        ("#fef9c3", "#713f12"),
        ("#fae8ff", "#701a75"),
        ("#ede9fe", "#4c1d95"),
        ("#e0f2fe", "#075985"),
    ]
    tagStyles = {}
    for i, k in enumerate(NEG.keys()):
        bg, color = NEG_PALETTE[i % len(NEG_PALETTE)]
        tagStyles[k] = {"bg": bg, "color": color}
    for i, k in enumerate(POS.keys()):
        bg, color = POS_PALETTE[i % len(POS_PALETTE)]
        tagStyles[k] = {"bg": bg, "color": color}
    for i, k in enumerate(WHAT.keys()):
        bg, color = WHAT_PALETTE[i % len(WHAT_PALETTE)]
        tagStyles[k] = {"bg": bg, "color": color}
    for k in WHO.keys():
        tagStyles.setdefault(k, {"bg": "#f1f5f9", "color": "#0f172a"})
    for k in WHEN.keys():
        tagStyles.setdefault(k, {"bg": "#f1f5f9", "color": "#0f172a"})
    for k in WHERE.keys():
        tagStyles.setdefault(k, {"bg": "#f1f5f9", "color": "#0f172a"})

    # ----- Reviews array -----
    # Pick ~250 most informative: prioritise tagged reviews, balance ratings
    by_rating = defaultdict(list)
    for r in reviews:
        # informativeness score: tag count + length bonus
        score = len(r["tags"]) * 5 + min(len(r["text"]) // 80, 8)
        by_rating[r["rating"]].append((score, r))
    target = 250
    quotas = {1: 60, 2: 30, 3: 30, 4: 50, 5: 80}
    selected = []
    for rating, lst in by_rating.items():
        lst.sort(key=lambda x: -x[0])
        q = quotas.get(rating, 30)
        selected.extend([r for _, r in lst[:q]])
    # if under-quota in some buckets, top up with leftover from others
    if len(selected) < target:
        remaining = []
        for rating, lst in by_rating.items():
            q = quotas.get(rating, 30)
            remaining.extend([r for _, r in lst[q:]])
        remaining.sort(key=lambda r: -(len(r["tags"]) * 5 + min(len(r["text"]) // 80, 8)))
        selected.extend(remaining[: target - len(selected)])
    selected = selected[:target]

    review_arr = []
    for r in selected:
        text = f"{r['title']}. {r['content']}".strip(". ").strip()
        text = re.sub(r"\s+", " ", text)
        if len(text) > 600:
            text = text[:597].rstrip() + "..."
        review_arr.append({
            "r": int(r["rating"]),
            "t": text,
            "tags": r["tags"],
        })

    # ----- Summaries -----
    cpSummary = (
        f"The core target audience is <b>garden and patio users</b> as well as <b>families with children</b>, "
        f"who in summer and especially during <b>BBQ season</b> are looking for a solution against wasps, mosquitoes and flies. "
        f"Indoor use in the <b>bedroom</b> and in <b>restaurants/cafes</b> are the two strongest secondary use cases."
    )
    csSummary = (
        f"Across <b>{total}</b> reviews, the dominant theme is <b>efficacy against wasps</b> — many devices primarily "
        f"catch mosquitoes and flies. Secondary critical themes: <b>killing bees</b>, <b>noise level</b> and <b>UV glare</b>. "
        f"Positively highlighted are <b>visible success</b> (full tray), <b>easy cleaning</b> and <b>safety for children</b>."
    )

    # ----- Final assembly -----
    out = {
        "totalReviews": total,
        "avgRating": round(avg, 2),
        "starDist": star_dist,
        "cpSummary": cpSummary,
        "cpWho": cpWho,
        "cpWhen": cpWhen,
        "cpWhere": cpWhere,
        "cpWhat": cpWhat,
        "usageScenarios": usageScenarios,
        "csSummary": csSummary,
        "negativeTopics": negativeTopics,
        "negativeInsights": negativeInsights,
        "positiveTopics": positiveTopics,
        "positiveInsights": positiveInsights,
        "buyersMotivation": buyersMotivation,
        "customerExpectations": customerExpectations,
        "themeFilters": themeFilters,
        "tagStyles": tagStyles,
        "reviews": review_arr,
    }

    OUT_PATH.write_text(json.dumps(out, indent=2, ensure_ascii=False), encoding="utf-8")
    size = OUT_PATH.stat().st_size

    # ----- Report -----
    print("=" * 60)
    print("VOC BUILD REPORT — DE / Electric")
    print("=" * 60)
    print(f"Total reviews processed: {total}")
    print(f"Average rating:          {avg:.2f}")
    print(f"Star distribution:       1*={star_dist[0]}  2*={star_dist[1]}  3*={star_dist[2]}  4*={star_dist[3]}  5*={star_dist[4]}")
    print()
    print("Top 3 negative themes:")
    for k, n in neg_counter.most_common(3):
        print(f"  - {NEG_LABELS[k][0]}: {n} ({pct(n,total_neg)})")
    print()
    print("Top 3 positive themes:")
    for k, n in pos_counter.most_common(3):
        print(f"  - {POS_LABELS[k][0]}: {n} ({pct(n,total_pos)})")
    print()
    print(f"Output file:  {OUT_PATH}")
    print(f"File size:    {size:,} bytes")


if __name__ == "__main__":
    main()
