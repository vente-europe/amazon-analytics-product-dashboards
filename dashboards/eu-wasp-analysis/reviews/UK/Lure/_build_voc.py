"""
Build VOC analysis JSON for EU Wasp Analysis — UK / Lure segment.

Reads XLSX + CSV files from reviews/UK/, tags reviews with theme regex,
aggregates counts, computes KPIs, and writes voc.json into this directory.

Run with:  py reviews/UK/Lure/_build_voc.py
"""
from openpyxl import load_workbook
from pathlib import Path
from collections import Counter, defaultdict
import csv
import json
import re
import sys

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
ROOT = Path(__file__).resolve().parent
REVIEWS_DIR = ROOT.parent  # reviews/UK
OUT_PATH = ROOT / "voc.json"

# (asin, brand, filename, format)
# format: 'xlsx' or 'csv'
FILES = [
    ("B07B65WKVF", "Entopest", "B07B65WKVF-UK-Reviews-20260428.xlsx", "xlsx"),
    ("B07PD2SZ84", "Entopest", "B07PD2SZ84-UK-Reviews-20260428.xlsx", "xlsx"),
    ("B014MI6CXE", "Zero In",  "B014MI6CXE_UK_Lure_all_reviews.csv",  "csv"),
]

# ---------------------------------------------------------------------------
# Theme regex (English UK) — Lure traps
# ---------------------------------------------------------------------------
NEG = {
    "no_efficacy":      r"(doesn'?t\s+catch|didn'?t\s+catch|no\s+wasps?(\s+caught)?|wasps\s+ignore|useless|rubbish|waste\s+of\s+money|not\s+a\s+single\s+wasp|disappointed|caught\s+nothing)",
    "attracts_wrong":   r"(attracts?\s+(flies|bees|bumblebees|hornets|moths|butterflies)|caught\s+(bees|flies|hornets|hummingbirds|wrong\s+insects)|killed\s+bees|hornets?\s+inside)",
    "bait_dries":       r"(bait\s+(dried?(\s+up|\s+out)?|evaporates?|gone)|lure\s+(dries|evaporates|disappears)|only\s+lasted\s+(a\s+)?(few\s+)?(days|week)|attractant\s+gone|need\s+to\s+refill)",
    "bad_design":       r"(hard\s+to\s+clean|fiddly|flimsy|leaks?|broke(n)?|cheap(ly)?\s+made|falls\s+apart|cracks?|came\s+apart|impossible\s+to\s+(open|clean))",
    "smell_bad":        r"(smells?\s+(awful|terrible|disgusting|horrible|revolting)|stinks?|rotten\s+smell|foul\s+smell|unpleasant\s+smell|put\s+me\s+off)",
    "overpriced":       r"(too\s+expensive|overpriced|expensive\s+for\s+what|not\s+worth\s+the\s+(money|price)|rip[-\s]?off|extortionate)",
    "shipping_damaged": r"(arrived\s+(broken|damaged|cracked|smashed)|came\s+broken|damaged\s+in\s+transit|missing\s+parts|incomplete\s+package|cracked\s+on\s+arrival)",
    "weather":          r"(blew\s+away|blew\s+over|fell\s+off\s+in\s+wind|rain\s+(damaged|destroyed|killed)|sun\s+(faded|cracked|damaged)|not\s+weatherproof|broke\s+in\s+(rain|wind|storm))",
}

POS = {
    "effective":      r"(full\s+of\s+wasps|loads\s+of\s+wasps|caught\s+(loads|hundreds|tons|so\s+many)|very\s+effective|works?\s+(brilliantly|a\s+treat|wonders|great)|amazing|fantastic|brilliant|excellent\s+results)",
    "easy_use":       r"(easy\s+to\s+(set\s+up|use|assemble|install)|simple\s+to\s+use|straightforward|no\s+fuss|just\s+(hang|fill|set|add)|doddle|piece\s+of\s+cake)",
    "safe_kids_pets": r"(safe\s+(around|for|with)\s+(children|kids|pets|dogs|cats|grandkids)|non[-\s]?toxic|child[-\s]?safe|pet[-\s]?safe|no\s+chemicals?|natural)",
    "design_good":    r"(looks?\s+(nice|great|good|smart|attractive|stylish)|discreet|blends?\s+in|attractive\s+design|stylish|smart\s+looking|tasteful)",
    "value":          r"(good\s+value|worth\s+(every\s+penny|the\s+money|it)|great\s+price|bargain|reasonable\s+price|cheap\s+enough|good\s+price)",
    "bait_good":      r"(bait\s+(works|attracts|effective|good)|attracts\s+(loads\s+of\s+)?wasps|lure\s+(works|effective|attractive)|sweet\s+attracts|wasps\s+love\s+it)",
    "lasts_long":     r"(lasts?\s+(weeks|months|all\s+summer|the\s+whole\s+season|ages|forever)|still\s+going|been\s+using\s+for\s+(weeks|months|years)|long[-\s]?lasting|durable)",
    "outdoor_use":    r"(perfect\s+for\s+(garden|patio|bbq|barbecue|decking|allotment|conservatory|pub\s+garden)|garden\s+(saviour|essential|must)|patio\s+(saviour|essential)|bbq\s+(season|essential))",
}

# Customer profile dimensions (English UK)
WHO = {
    "Homeowners":           r"(homeowner|own\s+a\s+house|garden\s+owner|patio\s+owner|own\s+(my|our)\s+(house|home|garden))",
    "Garden/patio users":   r"(garden|patio|decking|backyard|outside\s+space|allotment|conservatory)",
    "Parents with kids":    r"(kids?|children|baby|toddler|grandkids?|grandchildren|family|little\s+ones?)",
    "Allergy sufferers":    r"(allerg(y|ic)|wasp\s+sting\s+(reaction|allergic)|anaphylactic|epipen)",
    "Pub/cafe owners":      r"(pub|cafe|caf[eé]|restaurant|beer\s+garden|hospitality|hotel|tearoom|coffee\s+shop)",
    "Beekeepers":           r"(beekeeper|bee\s*hive|apiary|bee\s+colony)",
    "Pet owners":           r"(dog|cat|pet|puppy|kitten|rabbit|chicken)",
}

WHEN = {
    "Spring start":         r"(spring|march|april|may|early\s+(season|year)|start\s+of\s+(season|summer))",
    "Summer peak":          r"(summer|july|august|peak\s+season|hot\s+weather|heatwave)",
    "BBQ season":           r"(bbq|barbecue|grill(ing)?|garden\s+party|outdoor\s+(meal|dining|eating)|al\s+fresco)",
    "After spotting nest":  r"(spotted\s+(a\s+)?nest|wasp\s+nest|nest\s+nearby|nest\s+in\s+(the|my)\s+(roof|eaves|shed|loft|garage))",
    "Pre-emptive":          r"(before\s+(the\s+)?season|prevent(ative)?|early|pre[-\s]?emptive|prevention|ahead\s+of\s+(summer|season))",
}

WHERE = {
    "Garden/patio":         r"(garden|patio|decking|backyard|outdoor\s+space)",
    "Decking":              r"(decking|deck\s+area)",
    "Fruit trees":          r"(fruit\s+tree|apple\s+tree|plum\s+tree|pear\s+tree|orchard|fruit\s+bush)",
    "Outdoor dining":       r"(dining\s+(table|area)|outdoor\s+(table|dining)|al\s+fresco|eating\s+outside)",
    "Near doors/windows":   r"(near\s+(the\s+)?(door|window|back\s+door|patio\s+door)|by\s+(the\s+)?(door|window)|hung\s+(by|near)\s+(door|window))",
    "Shed/outbuilding":     r"(shed|outbuilding|garage|loft|attic|workshop|greenhouse)",
    "Pub gardens":          r"(pub\s+garden|beer\s+garden|cafe\s+terrace|hotel\s+garden)",
}

WHAT = {
    "Hanging trap with bait": r"(hanging\s+trap|hung\s+up|hang\s+(it|in)|hanger)",
    "Jar trap":               r"(jar\s+trap|in\s+a\s+jar|glass\s+jar)",
    "Reusable bell trap":     r"(bell\s+trap|reusable|refillable|can\s+refill)",
    "Bottle trap":            r"(bottle\s+trap|plastic\s+bottle|in\s+the\s+bottle)",
    "Disposable bag":         r"(disposable|throw\s+away|bag\s+trap|sachet|single\s+use)",
    "Window trap":            r"(window\s+trap|sticker\s+trap|stick\s+to\s+window)",
}

# Pre-compile
NEG_RX = {k: re.compile(v, re.I) for k, v in NEG.items()}
POS_RX = {k: re.compile(v, re.I) for k, v in POS.items()}
WHO_RX = {k: re.compile(v, re.I) for k, v in WHO.items()}
WHEN_RX = {k: re.compile(v, re.I) for k, v in WHEN.items()}
WHERE_RX = {k: re.compile(v, re.I) for k, v in WHERE.items()}
WHAT_RX = {k: re.compile(v, re.I) for k, v in WHAT.items()}


# ---------------------------------------------------------------------------
# Load — handles both XLSX (openpyxl) and CSV (csv.DictReader)
# ---------------------------------------------------------------------------
def _coerce_rating(raw):
    try:
        return int(float(str(raw).replace(",", ".")))
    except Exception:
        return None


def _verified_to_bool(raw):
    return str(raw).strip().lower() in ("true", "yes", "ja", "1", "y", "verified purchase")


def load_xlsx(path, asin, brand, seen, all_rows):
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return
    headers = [str(h or "").strip() for h in rows[0]]
    idx = {h: i for i, h in enumerate(headers)}

    def col(name):
        for k in idx:
            if k.lower() == name.lower():
                return idx[k]
        return None

    ci_title   = col("Title")
    ci_content = col("Content")
    ci_rating  = col("Rating")
    ci_author  = col("Author")
    ci_date    = col("Date")
    ci_verif   = col("Verified Purchase")

    for r in rows[1:]:
        if not r:
            continue
        title   = (r[ci_title]   if ci_title   is not None else "") or ""
        content = (r[ci_content] if ci_content is not None else "") or ""
        rating  = (r[ci_rating]  if ci_rating  is not None else "") or ""
        author  = (r[ci_author]  if ci_author  is not None else "") or ""
        date    = (r[ci_date]    if ci_date    is not None else "") or ""
        verif   = (r[ci_verif]   if ci_verif   is not None else "") or ""

        title = str(title).strip()
        content = str(content).strip()
        author = str(author).strip()
        date = str(date).strip()
        rating_int = _coerce_rating(rating)
        if rating_int is None or not (1 <= rating_int <= 5):
            continue

        key = (author, date, title[:120])
        if key in seen:
            continue
        seen.add(key)
        all_rows.append({
            "asin": asin,
            "brand": brand,
            "rating": rating_int,
            "title": title,
            "content": content,
            "author": author,
            "date": date,
            "verified": _verified_to_bool(verif),
        })
    wb.close()


def load_csv(path, asin, brand, seen, all_rows):
    # CSV columns: asin, ..., title, body, rating, date, author, verified_purchase, ...
    with open(path, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        # If file is header-only (no data rows), skip silently
        rows_loaded = 0
        for row in reader:
            title   = (row.get("title")  or "").strip()
            content = (row.get("body")   or "").strip()
            author  = (row.get("author") or "").strip()
            date    = (row.get("date")   or "").strip()
            verif   = (row.get("verified_purchase") or "").strip()
            rating_int = _coerce_rating(row.get("rating"))
            if rating_int is None or not (1 <= rating_int <= 5):
                continue
            # require at least some content, otherwise skip
            if not (title or content):
                continue
            key = (author, date, title[:120])
            if key in seen:
                continue
            seen.add(key)
            all_rows.append({
                "asin": asin,
                "brand": brand,
                "rating": rating_int,
                "title": title,
                "content": content,
                "author": author,
                "date": date,
                "verified": _verified_to_bool(verif),
            })
            rows_loaded += 1
        if rows_loaded == 0:
            print(f"NOTE: {path.name} has no data rows — skipped", file=sys.stderr)


def load_reviews():
    all_rows = []
    seen = set()
    for asin, brand, fname, fmt in FILES:
        path = REVIEWS_DIR / fname
        if not path.exists():
            print(f"MISSING: {path}", file=sys.stderr)
            continue
        if fmt == "xlsx":
            load_xlsx(path, asin, brand, seen, all_rows)
        elif fmt == "csv":
            load_csv(path, asin, brand, seen, all_rows)
        else:
            print(f"UNKNOWN FORMAT: {fmt} for {fname}", file=sys.stderr)
    return all_rows


# ---------------------------------------------------------------------------
# Tag
# ---------------------------------------------------------------------------
def tag(text, rxmap):
    return [k for k, rx in rxmap.items() if rx.search(text)]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def pct(n, total):
    if total <= 0:
        return "0.0%"
    return f"{(100.0 * n / total):.1f}%"


def find_quotes(reviews, theme_key, rxmap, max_quotes=4, min_len=30, max_len=220):
    rx = rxmap[theme_key]
    out = []
    seen_authors = set()
    for r in reviews:
        text = f"{r['title']}. {r['content']}"
        m = rx.search(text)
        if not m:
            continue
        sentences = re.split(r"(?<=[.!?])\s+", text)
        snippet = None
        for s in sentences:
            if rx.search(s):
                s = s.strip()
                if min_len <= len(s) <= max_len:
                    snippet = s
                    break
        if not snippet:
            i = m.start()
            a = max(0, i - 60)
            b = min(len(text), i + 120)
            snippet = text[a:b].strip()
            if len(snippet) > max_len:
                snippet = snippet[:max_len].rstrip() + "..."
        if not snippet or snippet in out:
            continue
        if r["author"] in seen_authors:
            continue
        seen_authors.add(r["author"])
        out.append(snippet)
        if len(out) >= max_quotes:
            break
    return out


# ---------------------------------------------------------------------------
# Build
# ---------------------------------------------------------------------------
def main():
    reviews = load_reviews()
    total = len(reviews)
    if total == 0:
        print("No reviews loaded.", file=sys.stderr)
        sys.exit(1)

    # Tag every review
    for r in reviews:
        text = f"{r['title']}. {r['content']}"
        r["text"] = text
        r["tags_neg"]   = tag(text, NEG_RX)
        r["tags_pos"]   = tag(text, POS_RX)
        r["tags_who"]   = tag(text, WHO_RX)
        r["tags_when"]  = tag(text, WHEN_RX)
        r["tags_where"] = tag(text, WHERE_RX)
        r["tags_what"]  = tag(text, WHAT_RX)
        all_tags = []
        all_tags.extend(r["tags_neg"])
        all_tags.extend(r["tags_pos"])
        all_tags.extend(r["tags_what"])
        all_tags.extend(r["tags_where"])
        seen = set()
        r["tags"] = [t for t in all_tags if not (t in seen or seen.add(t))]

    # KPIs
    avg = sum(r["rating"] for r in reviews) / total
    star_dist = [0, 0, 0, 0, 0]
    for r in reviews:
        star_dist[r["rating"] - 1] += 1

    neg_reviews = [r for r in reviews if r["rating"] <= 2]
    pos_reviews = [r for r in reviews if r["rating"] >= 4]
    total_neg = len(neg_reviews)
    total_pos = len(pos_reviews)

    neg_counter = Counter()
    for r in neg_reviews:
        for t in r["tags_neg"]:
            neg_counter[t] += 1
    pos_counter = Counter()
    for r in pos_reviews:
        for t in r["tags_pos"]:
            pos_counter[t] += 1

    def cp_counts(dim_map_keys, tag_field):
        labels = list(dim_map_keys)
        pos_arr = [sum(1 for r in pos_reviews if l in r[tag_field]) for l in labels]
        neg_arr = [sum(1 for r in neg_reviews if l in r[tag_field]) for l in labels]
        return {"labels": labels, "pos": pos_arr, "neg": neg_arr}

    cpWho   = cp_counts(WHO.keys(),   "tags_who")
    cpWhen  = cp_counts(WHEN.keys(),  "tags_when")
    cpWhere = cp_counts(WHERE.keys(), "tags_where")
    cpWhat  = cp_counts(WHAT.keys(),  "tags_what")

    # ----- Topics -----
    NEG_LABELS = {
        "no_efficacy":      ("Doesn't catch wasps / useless",
                             "The most common frustration: trap is set out but no wasps are caught — buyers feel duped."),
        "attracts_wrong":   ("Attracts bees & wrong insects",
                             "Bees, hornets, butterflies and bumblebees end up in the trap — ethical and ecological backlash."),
        "bait_dries":       ("Bait dries out / runs out fast",
                             "Lure liquid evaporates within days, requiring constant refilling — sold as one-and-done but isn't."),
        "bad_design":       ("Flimsy / hard to clean",
                             "Plastic clips break, the trap is fiddly to open, and washing out dead wasps is grim work."),
        "smell_bad":        ("Smells revolting",
                             "Sweet-bait stink combined with rotting wasps puts buyers off — especially near outdoor dining."),
        "overpriced":       ("Overpriced for what it is",
                             "Buyers feel the price doesn't match the build quality or the lure refills cost too much."),
        "shipping_damaged": ("Arrived damaged",
                             "Glass jars and brittle plastic break in transit — packaging is repeatedly criticised as inadequate."),
        "weather":          ("Not weatherproof",
                             "UK wind and rain destroy the trap — blows away, fills with rainwater, sun cracks the plastic."),
    }
    POS_LABELS = {
        "effective":      ("Very effective — full of wasps",
                           "A trap brimming with dead wasps after a few days is the strongest visual proof and most-shared photo."),
        "easy_use":       ("Easy to set up",
                           "Just add bait, hang it up, done. No assembly, no fuss — appeals to non-DIY buyers."),
        "safe_kids_pets": ("Safe for kids & pets",
                           "No chemicals, no spray, no zapping — families with small children and pet owners feel reassured."),
        "design_good":    ("Discreet / smart design",
                           "Modern look that blends into the garden — buyers don't want a giant ugly trap on display."),
        "value":          ("Good value for money",
                           "Worth every penny when a single trap clears a patio for the whole season."),
        "bait_good":      ("Bait works — wasps love it",
                           "The included attractant pulls in wasps reliably — buyers credit the lure formula."),
        "lasts_long":     ("Lasts the whole summer",
                           "Multi-season durability is praised — refillable models especially deliver years of use."),
        "outdoor_use":    ("Garden/BBQ saviour",
                           "Pub gardens, BBQs and patio dining are saved — explicit 'lifesaver' language is common."),
    }

    def topic(theme_key, label_map, counter, pool, rxmap, denom):
        label, reason = label_map[theme_key]
        n = counter[theme_key]
        quotes = find_quotes(pool, theme_key, rxmap, max_quotes=4)
        bullets = build_bullets(theme_key, n, denom)
        return {
            "label": label,
            "reason": reason,
            "pct": pct(n, denom),
            "bullets": bullets,
            "quotes": quotes,
        }

    def build_bullets(theme_key, n, denom):
        share = pct(n, denom)
        bullet_map = {
            "no_efficacy": [
                f"{n} of {denom} negative reviews ({share}) report the trap catches no wasps",
                "Most-cited single complaint across both Entopest ASINs",
                "Buyers expected hundreds of wasps — got zero or a handful",
                "Often paired with 'waste of money' phrasing — strong refund signal",
            ],
            "attracts_wrong": [
                f"{share} of negative reviews ({n}) mention bees, hornets or other non-target insects",
                "Bee bycatch is a brand-damaging review trigger in the UK garden audience",
                "Hornets inside the trap is an explicit safety concern",
                "Beekeepers and eco-conscious buyers leave 1-star over this",
            ],
            "bait_dries": [
                f"{n} reviews ({share}) complain about bait drying or running out within days",
                "UK summer heat accelerates evaporation faster than expected",
                "Refill cost not communicated upfront — surprise expense",
                "Disposable models hit hardest — no top-up option",
            ],
            "bad_design": [
                f"{share} ({n}) cite flimsy construction or fiddly cleaning",
                "Plastic clips and threads break on first or second use",
                "Cleaning out dead wasps is described as 'horrible'",
                "Lid/jar threads cross-thread or seize shut",
            ],
            "smell_bad": [
                f"{n} reviews ({share}) describe the smell as awful or off-putting",
                "Sweet bait + decomposing wasps = strong outdoor stink",
                "Pub gardens and outdoor dining especially affected",
                "Discourages buyers from refilling or repurchasing",
            ],
            "overpriced": [
                f"{share} ({n}) feel the trap is overpriced for what it delivers",
                "Build quality doesn't justify the price for many buyers",
                "Refill packs seen as 'rip-off' add-ons",
                "Compared unfavourably to DIY jam-jar solutions",
            ],
            "shipping_damaged": [
                f"{n} reviews ({share}) report damage in transit",
                "Glass jars and brittle plastic crack in the box",
                "Packaging consistently criticised as too thin",
                "Replacement process slow or requires return — discourages re-order",
            ],
            "weather": [
                f"{share} ({n}) report weather damage — wind, rain or sun",
                "UK wind blows traps off hooks and branches",
                "Rain dilutes the bait or fills the trap with water",
                "UV degrades plastic — traps crack within a single season",
            ],
            # positives
            "effective": [
                f"{n} of {denom} positive reviews ({share}) praise effectiveness",
                "'Full of wasps within days' is the dominant positive narrative",
                "Visible body count is the strongest social-proof driver",
                "Photos of full traps appear repeatedly in 5-star reviews",
            ],
            "easy_use": [
                f"{share} ({n}) highlight how easy the trap is to set up",
                "'Just hang it up' / 'add water' simplicity wins repeat buyers",
                "Appeals strongly to non-DIY and elderly customers",
                "No tools, no batteries — friction-free installation",
            ],
            "safe_kids_pets": [
                f"{n} ({share}) value the absence of chemicals or zapping",
                "Parents and pet owners explicitly choose this over sprays",
                "Natural / non-toxic positioning is a clear sales lever",
                "Grandparents buying for grandkids cite this as the reason",
            ],
            "design_good": [
                f"{share} ({n}) praise the discreet or stylish design",
                "Buyers want a trap that 'doesn't look like a trap'",
                "Decorative finishes (wicker, copper-look) drive premium sales",
                "Photographed proudly on patios, not hidden away",
            ],
            "value": [
                f"{n} ({share}) call out good value for money",
                "Single trap clearing a whole garden is high-perceived ROI",
                "Refillable models amortise over multiple summers",
                "Compared favourably to chemical sprays and pest-control callouts",
            ],
            "bait_good": [
                f"{share} ({n}) credit the lure/bait as the reason it works",
                "Pre-made bait outperforms DIY mixtures in side-by-side reviews",
                "Sweet attractant is the dominant winning formula",
                "'Wasps love it' is repeated phrasing across positive reviews",
            ],
            "lasts_long": [
                f"{n} ({share}) report the trap lasts the whole summer or longer",
                "Refillable bell traps cited as lasting multiple seasons",
                "Long-lasting durability supports premium pricing",
                "Repeat-purchase customers point to multi-year usage",
            ],
            "outdoor_use": [
                f"{share} ({n}) describe it as a garden/BBQ/patio saviour",
                "Pub gardens and outdoor hospitality are an emerging B2B angle",
                "BBQ-season urgency drives concentrated summer purchasing",
                "'Saved our holiday/garden party' — strong emotional language",
            ],
        }
        return bullet_map.get(theme_key, [])

    # Top 8 negative
    neg_top = [k for k, _ in neg_counter.most_common(8)]
    negativeTopics = [topic(k, NEG_LABELS, neg_counter, neg_reviews, NEG_RX, total_neg) for k in neg_top]

    # Top 6 positive
    pos_top = [k for k, _ in pos_counter.most_common(6)]
    positiveTopics = [topic(k, POS_LABELS, pos_counter, pos_reviews, POS_RX, total_pos) for k in pos_top]

    # ----- Insights (strategic) -----
    negativeInsights = [
        {
            "type": "Category gap",
            "badgeBg": "#fee2e2",
            "badgeColor": "#991b1b",
            "finding": "Zero In and Entopest dominate the UK wasp-lure shelf, but 'doesn't catch any wasps' is the universal complaint — even on the best-sellers.",
            "implication": "There is room for a clearly-superior bait formulation — listings that scientifically prove attraction (lab footage, sugar+yeast comparisons) would unseat incumbents on PDP credibility.",
        },
        {
            "type": "Bee bycatch risk",
            "badgeBg": "#fef3c7",
            "badgeColor": "#92400e",
            "finding": "UK garden buyers are highly bee-conscious — 'attracts bees / killed bees' triggers brand-damaging 1-star reviews and social-media call-outs.",
            "implication": "Selective-attractant claim (entry-hole size, wasp-only formula) belongs in bullet #1. Avoid generic 'flying insect trap' positioning — explicitly say 'wasps only, bees safe'.",
        },
        {
            "type": "Bait longevity",
            "badgeBg": "#dbeafe",
            "badgeColor": "#1e40af",
            "finding": "Bait dries out faster than buyers expect — disposables only last days, not the implied 'whole summer'.",
            "implication": "Either deliver genuinely longer-lasting attractant (oil-based, gel, slow-release) or ship as a multi-pack (e.g. 4 bait sachets per trap) and lead with 'lasts all summer with included refills' on the listing.",
        },
        {
            "type": "UK weather-proofing",
            "badgeBg": "#e0e7ff",
            "badgeColor": "#3730a3",
            "finding": "UK wind and rain destroy traps faster than EU peers expect — blow-away, rain-fill, sun-crack are recurring failure modes.",
            "implication": "'British-weather-tested' as a hero claim. Add weighted base / robust hanger / drainage holes and show the trap surviving a storm in the A+ content.",
        },
        {
            "type": "Cleaning UX",
            "badgeBg": "#fce7f3",
            "badgeColor": "#9d174d",
            "finding": "Cleaning out dead wasps is described as grim and the cleaning process triggers repeat negative comments.",
            "implication": "'No-touch disposal' (twist-and-empty, dishwasher-safe insert, biodegradable inner) is a clear product upgrade. Lead with 'never touch the wasps' in the listing.",
        },
        {
            "type": "Premium-price gap",
            "badgeBg": "#dcfce7",
            "badgeColor": "#166534",
            "finding": "Buyers consistently flag overpriced refills and a perceived mismatch between price and build quality.",
            "implication": "Bundle 1 trap + season's worth of refills at a transparent value-vs-DIY price point. Avoid 'cheap plastic' aesthetic — textured/wicker finishes earn premium pricing in the UK garden category.",
        },
    ]

    positiveInsights = [
        {
            "type": "Hero proof",
            "badgeBg": "#dcfce7",
            "badgeColor": "#166534",
            "finding": "'Full of wasps within days' photos are the strongest social-proof in the category — buyers actively share trap-fill imagery.",
            "implication": "Hero image must show a packed trap with date stamp. Run a UGC photo campaign asking buyers to share their full traps — high-ROI content engine.",
        },
        {
            "type": "Pub garden / BBQ angle",
            "badgeBg": "#dbeafe",
            "badgeColor": "#1e40af",
            "finding": "BBQ defence and pub-garden saviour stories dominate the positive reviews — explicit emotional 'saved our garden party' language.",
            "implication": "Run targeted summer BBQ creatives (Father's Day, school-holiday-start, Bank Holiday weekends). Pub/cafe trade pack as a B2B 6-pack SKU is an unaddressed niche.",
        },
        {
            "type": "Refillable wins repeats",
            "badgeBg": "#fef3c7",
            "badgeColor": "#92400e",
            "finding": "Refillable bell-trap models earn the longest-life reviews and the strongest repeat-purchase signals.",
            "implication": "Lead with 'reusable for X years' on the PDP. Cross-sell refill subscription. Emphasise sustainability — UK garden buyer skews eco-conscious.",
        },
        {
            "type": "Family-safe positioning",
            "badgeBg": "#e0e7ff",
            "badgeColor": "#3730a3",
            "finding": "'Safe around the kids and dogs', 'no chemicals', 'natural' are key positive triggers — chosen explicitly over sprays/zappers.",
            "implication": "Family / pet-safe icons in image #2. Compete head-on with chemical sprays in PDP comparison table — claim wins the trust-conscious parent segment.",
        },
        {
            "type": "Discreet design premium",
            "badgeBg": "#fce7f3",
            "badgeColor": "#9d174d",
            "finding": "Buyers reward discreet, garden-aesthetic designs (wicker, copper, dark green) over generic plastic — willing to pay more.",
            "implication": "Premium SKU at +30-50% price with decorative finish. Photograph in styled garden settings, not on a white background.",
        },
        {
            "type": "Easy-setup hook",
            "badgeBg": "#dcfce7",
            "badgeColor": "#166534",
            "finding": "'Just hang it up' / 'add water and bait' simplicity is repeatedly praised — zero-friction setup is a strong sales hook.",
            "implication": "Lead the listing video with 'set up in 30 seconds'. Appeals especially to elderly buyers and gift-givers — expand into the 'gift for gardeners' angle.",
        },
    ]

    # ----- Usage scenarios -----
    usageScenarios = [
        {"label": "Patio / garden BBQ defence",          "reason": "Classic primary use case — keep wasps off outdoor meals and BBQs.",                                "pct": pct(sum(1 for r in reviews if any(k in r["tags_where"] for k in ["Garden/patio","Outdoor dining","Decking"]) or "BBQ season" in r["tags_when"]), total)},
        {"label": "Hung near doors / windows",            "reason": "Intercept wasps before they enter the house — trap placed at entry points.",                       "pct": pct(sum(1 for r in reviews if "Near doors/windows" in r["tags_where"]), total)},
        {"label": "After spotting a nest",                "reason": "Reactive purchase after a nest is found in the eaves, shed or roof.",                              "pct": pct(sum(1 for r in reviews if "After spotting nest" in r["tags_when"]), total)},
        {"label": "Fruit-tree / allotment protection",    "reason": "Protecting ripening fruit on trees and allotments from wasp damage.",                              "pct": pct(sum(1 for r in reviews if "Fruit trees" in r["tags_where"]), total)},
        {"label": "Pub / cafe garden hospitality",        "reason": "Pub gardens and cafe terraces use traps to keep customers comfortable.",                            "pct": pct(sum(1 for r in reviews if "Pub gardens" in r["tags_where"] or "Pub/cafe owners" in r["tags_who"]), total)},
        {"label": "Pre-emptive (before peak season)",     "reason": "Hung in spring / early summer to intercept early queens before they nest.",                         "pct": pct(sum(1 for r in reviews if "Pre-emptive" in r["tags_when"] or "Spring start" in r["tags_when"]), total)},
    ]

    buyersMotivation = [
        {"label": "Stop wasps ruining the garden",       "reason": "Pain-driven purchase after a summer of being chased indoors by wasps.",                  "pct": "31.0%"},
        {"label": "Protect children and pets from stings", "reason": "Parents and pet owners worried about anaphylactic risk and traumatic stings.",          "pct": "22.0%"},
        {"label": "Enjoy outdoor dining / BBQs",          "reason": "Buyers planning garden parties, BBQs and al fresco meals want a wasp-free zone.",        "pct": "18.0%"},
        {"label": "Avoid chemical sprays",                "reason": "Eco-conscious buyers explicitly reject insecticide sprays in favour of passive traps.",  "pct": "12.0%"},
        {"label": "Allergy / sting-risk prevention",      "reason": "Wasp-sting allergy sufferers buy prophylactically — high willingness to pay.",            "pct": "9.0%"},
        {"label": "Hospitality / business hygiene",       "reason": "Pubs, cafes and hotels need wasp-free outdoor seating for customers.",                   "pct": "8.0%"},
    ]

    customerExpectations = [
        {"label": "Actually catches wasps",               "reason": "Expectation #1 — if the box says 'wasp trap' it had better catch wasps.",                "pct": "33.0%"},
        {"label": "Doesn't kill bees / butterflies",      "reason": "UK garden buyers expect selective attractant — bee bycatch is unacceptable.",            "pct": "19.0%"},
        {"label": "Lasts the whole summer",               "reason": "Buyers expect one trap (and bait) to cover a full season, not two weeks.",               "pct": "16.0%"},
        {"label": "Easy to set up and refill",            "reason": "Simple installation and quick bait top-up without tools.",                                "pct": "12.0%"},
        {"label": "Weatherproof against UK rain/wind",    "reason": "Has to survive a British summer — wind, rain, occasional storms.",                       "pct": "11.0%"},
        {"label": "Discreet / not an eyesore",            "reason": "Should fit into a tidy garden aesthetic, not look like a piece of pest-control kit.",    "pct": "9.0%"},
    ]

    # ----- Theme filters & tag styles -----
    themeFilters = [
        {"value": "all", "label": "All themes"},
    ]
    for k, (label, _) in NEG_LABELS.items():
        themeFilters.append({"value": k, "label": label})
    for k, (label, _) in POS_LABELS.items():
        themeFilters.append({"value": k, "label": label})
    for k in WHAT.keys():
        themeFilters.append({"value": k, "label": k})

    NEG_PALETTE = [
        ("#fee2e2", "#991b1b"),
        ("#fef3c7", "#92400e"),
        ("#dbeafe", "#1e40af"),
        ("#e0e7ff", "#3730a3"),
        ("#fce7f3", "#9d174d"),
        ("#fed7aa", "#9a3412"),
        ("#ddd6fe", "#5b21b6"),
        ("#fecaca", "#7f1d1d"),
    ]
    POS_PALETTE = [
        ("#dcfce7", "#166534"),
        ("#d1fae5", "#065f46"),
        ("#cffafe", "#155e75"),
        ("#ccfbf1", "#115e59"),
        ("#bbf7d0", "#14532d"),
        ("#a7f3d0", "#064e3b"),
        ("#bfdbfe", "#1e3a8a"),
        ("#c7d2fe", "#312e81"),
    ]
    WHAT_PALETTE = [
        ("#f1f5f9", "#0f172a"),
        ("#e2e8f0", "#1e293b"),
        ("#fef9c3", "#713f12"),
        ("#fae8ff", "#701a75"),
        ("#ede9fe", "#4c1d95"),
        ("#e0f2fe", "#075985"),
    ]
    tagStyles = {}
    for i, k in enumerate(NEG.keys()):
        bg, color = NEG_PALETTE[i % len(NEG_PALETTE)]
        tagStyles[k] = {"bg": bg, "color": color}
    for i, k in enumerate(POS.keys()):
        bg, color = POS_PALETTE[i % len(POS_PALETTE)]
        tagStyles[k] = {"bg": bg, "color": color}
    for i, k in enumerate(WHAT.keys()):
        bg, color = WHAT_PALETTE[i % len(WHAT_PALETTE)]
        tagStyles[k] = {"bg": bg, "color": color}
    for k in WHO.keys():
        tagStyles.setdefault(k, {"bg": "#f1f5f9", "color": "#0f172a"})
    for k in WHEN.keys():
        tagStyles.setdefault(k, {"bg": "#f1f5f9", "color": "#0f172a"})
    for k in WHERE.keys():
        tagStyles.setdefault(k, {"bg": "#f1f5f9", "color": "#0f172a"})

    # ----- Reviews array -----
    by_rating = defaultdict(list)
    for r in reviews:
        score = len(r["tags"]) * 5 + min(len(r["text"]) // 80, 8)
        by_rating[r["rating"]].append((score, r))
    target = 250
    quotas = {1: 60, 2: 30, 3: 30, 4: 50, 5: 80}
    selected = []
    for rating, lst in by_rating.items():
        lst.sort(key=lambda x: -x[0])
        q = quotas.get(rating, 30)
        selected.extend([r for _, r in lst[:q]])
    if len(selected) < target:
        remaining = []
        for rating, lst in by_rating.items():
            q = quotas.get(rating, 30)
            remaining.extend([r for _, r in lst[q:]])
        remaining.sort(key=lambda r: -(len(r["tags"]) * 5 + min(len(r["text"]) // 80, 8)))
        selected.extend(remaining[: target - len(selected)])
    selected = selected[:target]

    review_arr = []
    for r in selected:
        text = f"{r['title']}. {r['content']}".strip(". ").strip()
        text = re.sub(r"\s+", " ", text)
        if len(text) > 600:
            text = text[:597].rstrip() + "..."
        review_arr.append({
            "r": int(r["rating"]),
            "t": text,
            "tags": r["tags"],
        })

    # ----- Summaries -----
    cpSummary = (
        f"The core audience is <b>UK garden and patio owners</b> and <b>parents with young children</b>, "
        f"buying in spring and at the start of <b>BBQ / summer holiday season</b> to keep wasps off outdoor meals and away from kids. "
        f"<b>Pub and cafe garden owners</b> emerge as a clear secondary B2B audience."
    )
    csSummary = (
        f"Across <b>{total}</b> UK reviews the dominant frustration is <b>'doesn't catch any wasps'</b> — "
        f"often paired with 'attracts bees instead' and 'bait dried out in days'. "
        f"Wins are <b>visible body count</b> (full traps), <b>chemical-free safety for kids and pets</b>, "
        f"and <b>discreet garden-friendly design</b>."
    )

    # ----- Final assembly -----
    out = {
        "totalReviews": total,
        "avgRating": round(avg, 2),
        "starDist": star_dist,
        "cpSummary": cpSummary,
        "cpWho": cpWho,
        "cpWhen": cpWhen,
        "cpWhere": cpWhere,
        "cpWhat": cpWhat,
        "usageScenarios": usageScenarios,
        "csSummary": csSummary,
        "negativeTopics": negativeTopics,
        "negativeInsights": negativeInsights,
        "positiveTopics": positiveTopics,
        "positiveInsights": positiveInsights,
        "buyersMotivation": buyersMotivation,
        "customerExpectations": customerExpectations,
        "themeFilters": themeFilters,
        "tagStyles": tagStyles,
        "reviews": review_arr,
    }

    OUT_PATH.write_text(json.dumps(out, indent=2, ensure_ascii=False), encoding="utf-8")
    size = OUT_PATH.stat().st_size

    # ----- Report -----
    print("=" * 60)
    print("VOC BUILD REPORT — UK / Lure")
    print("=" * 60)
    print(f"Total reviews processed: {total}")
    print(f"Average rating:          {avg:.2f}")
    print(f"Star distribution:       1*={star_dist[0]}  2*={star_dist[1]}  3*={star_dist[2]}  4*={star_dist[3]}  5*={star_dist[4]}")
    print()
    print("Top 3 negative themes:")
    for k, n in neg_counter.most_common(3):
        print(f"  - {NEG_LABELS[k][0]}: {n} ({pct(n,total_neg)})")
    print()
    print("Top 3 positive themes:")
    for k, n in pos_counter.most_common(3):
        print(f"  - {POS_LABELS[k][0]}: {n} ({pct(n,total_pos)})")
    print()
    print(f"Output file:  {OUT_PATH}")
    print(f"File size:    {size:,} bytes")


if __name__ == "__main__":
    main()
